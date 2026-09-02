import io
from unittest.mock import patch

import frappe
from frappe.tests import UnitTestCase
from openpyxl import Workbook

from china_finance.overrides.bank_statement_import import get_full_import_preview, _validate_import_batch
from china_finance.services.bank_statement_import import parse_cmb_statement


class TestBankStatementImport(UnitTestCase):
	def test_cmb_statement_is_converted_to_native_bank_transaction_template(self):
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.append(["招商银行交易查询"])
		worksheet.append([])
		worksheet.append([
			"账号", "币种", "交易日", "交易类型", "借方金额", "贷方金额", "摘要", "流水号",
			"收(付)方名称", "收(付)方账号", "收(付)方开户行名",
		])
		worksheet.append([
			"769915033910001", "人民币", "2026-06-01", "提回对公户收款", None, 30000,
			"悦为智能投资款", "C0347G000073PQZ", "周悦", "6214867556395656", "招商银行",
		])
		worksheet.append([
			"769915033910001", "人民币", "2026-06-01", "对公转账正常提出", 24440, None,
			"租金", "C0347G0000HD0NZ", "东莞市润企产业运营服务有限公司", "44285901040013398",
			"中国农业银行股份有限公司东莞高新支行",
		])
		content = io.BytesIO()
		workbook.save(content)

		rows = parse_cmb_statement(content.getvalue(), "银行存款_悦为智能公户 - 招商银行")
		self.assertEqual(rows, [
			[
				"2026/6/1", "30000", "", "悦为智能投资款｜对方：周悦｜账号：6214867556395656｜开户行：招商银行｜交易类型：提回对公户收款",
				"C0347G000073PQZ", "银行存款_悦为智能公户 - 招商银行", "CNY",
			],
			[
				"2026/6/1", "", "24440", "租金｜对方：东莞市润企产业运营服务有限公司｜账号：44285901040013398｜开户行：中国农业银行股份有限公司东莞高新支行｜交易类型：对公转账正常提出",
				"C0347G0000HD0NZ", "银行存款_悦为智能公户 - 招商银行", "CNY",
			],
		])

	def test_cmb_statement_requires_native_headers(self):
		workbook = Workbook()
		workbook.active.append(["日期", "金额"])
		content = io.BytesIO()
		workbook.save(content)
		self.assertRaises(frappe.ValidationError, parse_cmb_statement, content.getvalue(), "测试银行账户")

	def test_cmb_statement_keeps_party_and_trace_metadata(self):
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.append([
			"账号", "币种", "交易日", "交易时间", "起息日", "交易类型", "借方金额", "贷方金额",
			"余额", "摘要", "流水号", "流程实例号", "业务参考号", "收(付)方名称", "收(付)方账号",
			"收(付)方开户行名", "收(付)方开户行地址", "扩展摘要", "交易分析码",
		])
		worksheet.append([
			"769915033910001", "人民币", "2026-07-01", "08:00:00", "2026-07-01", "对公转账正常提出",
			100, None, 900, "报销", "FLOW-001", "PROC-001", "BIZ-001", "张三", "6214",
			"招商银行", "东莞", "补充说明", "CPGATR",
		])
		content = io.BytesIO()
		workbook.save(content)

		rows = parse_cmb_statement(content.getvalue(), "测试银行账户")

		self.assertEqual(rows[0][4], "FLOW-001")
		self.assertIn("对方：张三", rows[0][3])
		self.assertIn("交易类型：对公转账正常提出", rows[0][3])
		self.assertIn("业务参考号：BIZ-001", rows[0][3])
		self.assertIn("流程实例号：PROC-001", rows[0][3])
		self.assertIn("开户地址：东莞", rows[0][3])
		self.assertIn("扩展摘要：补充说明", rows[0][3])
		self.assertIn("交易分析码：CPGATR", rows[0][3])

	def test_cmb_statement_rejects_broken_balance_sequence(self):
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.append(["账号", "币种", "交易日", "借方金额", "贷方金额", "余额", "摘要", "流水号"])
		worksheet.append(["769915033910001", "人民币", "2026-07-01", 100, None, 900, "费用", "FLOW-001"])
		worksheet.append(["769915033910001", "人民币", "2026-07-02", None, 50, 1000, "收款", "FLOW-002"])
		content = io.BytesIO()
		workbook.save(content)

		self.assertRaises(
			frappe.ValidationError,
			parse_cmb_statement,
			content.getvalue(),
			"测试银行账户",
		)

	def test_bank_statement_preview_returns_every_converted_row(self):
		class Column:
			def as_dict(self):
				return frappe._dict({
					"header_title": "日期",
					"df": frappe._dict({
						"fieldtype": "Date", "fieldname": "date", "label": "日期", "options": None,
						"parent": None, "reqd": 0, "default": None, "read_only": 0,
					}),
				})

		class Row:
			def __init__(self, row_number):
				self.row_number = row_number

			def as_list(self):
				return [f"2026/6/{self.row_number}"]

		class ImportFile:
			columns = [Column()]
			data = [Row(row_number) for row_number in range(2, 38)]

			def get_warnings(self):
				return []

		importer = frappe._dict(import_file=ImportFile(), data_import=frappe._dict(name="BSI-TEST"))
		with patch("china_finance.overrides.bank_statement_import.frappe.get_all", return_value=[]):
			preview = get_full_import_preview(importer)

		self.assertEqual(len(preview.data), 36)
		self.assertNotIn("max_rows_exceeded", preview)

	def test_import_batch_rejects_overlong_reference_before_writes(self):
		with self.assertRaises(frappe.ValidationError):
			_validate_import_batch([{"reference": "R" * 141}])

	def test_import_batch_rejects_duplicate_references_before_writes(self):
		with self.assertRaises(frappe.ValidationError):
			_validate_import_batch([{"reference": "FLOW-001"}, {"reference": "FLOW-001"}])
