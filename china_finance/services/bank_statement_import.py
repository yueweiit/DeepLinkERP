import csv
import hashlib
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import frappe
import openpyxl
from frappe import _
from frappe.utils.file_manager import save_file


STANDARD_HEADERS = ["日期", "存款", "取款", "摘要", "描述", "参考号码", "银行账户", "货币"]
SUPPORTED_BANK_PARSERS = {"招商银行": "cmb", "招商": "cmb"}
REQUIRED_CMB_HEADERS = {"交易日", "借方金额", "贷方金额", "摘要", "流水号"}
CMB_OPTIONAL_HEADERS = {"交易类型", "收(付)方名称", "收(付)方账号", "收(付)方开户行名"}
CURRENCY_MAP = {"人民币": "CNY", "CNY": "CNY", "美元": "USD", "USD": "USD", "港币": "HKD", "HKD": "HKD"}


@frappe.whitelist()
def convert_bank_statement(data_import, source_file, bank=None):
	"""Convert a supported bank statement to ERPNext's Bank Transaction template."""
	doc = frappe.get_doc("Bank Statement Import", data_import)
	doc.check_permission("write")
	_validate_import_context(doc, bank)
	file_doc = _get_attached_source_file(doc, source_file)

	parser = SUPPORTED_BANK_PARSERS.get(doc.bank)
	if not parser:
		frappe.throw(
			_("暂未支持银行 {0} 的流水格式，请使用原生“添加文件”上传标准模板。").format(doc.bank),
			title=_("银行流水转换"),
		)

	if parser != "cmb":
		frappe.throw(_("未找到银行 {0} 的流水解析器。").format(doc.bank))

	content = file_doc.get_content()
	content = content.encode() if isinstance(content, str) else content
	rows = parse_cmb_statement(content, doc.bank_account)
	file_hash = hashlib.sha256(content).hexdigest()
	converted_file, reused = _get_or_create_converted_file(doc, rows, file_hash)

	if doc.import_file != converted_file.file_url:
		doc.import_file = converted_file.file_url
		doc.save()

	return {
		"file_url": converted_file.file_url,
		"row_count": len(rows),
		"reused": reused,
	}


@frappe.whitelist()
def convert_bank_statement_import_log(statement_import_id, source_file):
	"""Convert a supported bank workbook for the new Banking importer.

	The React Banking importer stores uploads in ``Bank Statement Import Log``;
	keep the CMB parser here and replace the log's file with the normalized CSV
	that ERPNext already understands.  The original workbook remains attached
	to the same log for auditability.
	"""
	doc = frappe.get_doc("Bank Statement Import Log", statement_import_id)
	doc.check_permission("write")
	bank_account = frappe.db.get_value("Bank Account", doc.bank_account, ["bank", "account"] , as_dict=True)
	if not bank_account or bank_account.bank not in SUPPORTED_BANK_PARSERS:
		return {"supported": False}

	file_doc = _get_attached_source_file(doc, source_file, "Bank Statement Import Log")
	content = file_doc.get_content()
	content = content.encode() if isinstance(content, str) else content
	if Path(source_file).suffix.lower() != ".xlsx":
		frappe.throw(_("招商银行模板必须是 XLSX 文件。"), title=_("银行流水转换"))
	rows = parse_cmb_statement(content, doc.bank_account)
	file_hash = hashlib.sha256(content).hexdigest()
	converted_file, reused = _get_or_create_converted_file(
		doc, rows, file_hash, doctype="Bank Statement Import Log", fieldname="file"
	)
	doc.file = converted_file.file_url
	doc.save()
	return {"supported": True, "file_url": converted_file.file_url, "row_count": len(rows), "reused": reused}


def parse_cmb_statement(content, bank_account):
	"""Return normalized ERPNext Bank Transaction template rows from a CMB workbook."""
	try:
		workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=True)
	except Exception as exc:
		frappe.throw(_("无法读取招商银行 XLSX 流水：{0}").format(exc), title=_("银行流水转换"))

	for worksheet in workbook.worksheets:
		header_row, header_indexes = _find_cmb_header(worksheet)
		if header_row:
			return _parse_cmb_rows(worksheet, header_row, header_indexes, bank_account)

	frappe.throw(
		_("未识别到招商银行流水表头，缺少：{0}").format("、".join(sorted(REQUIRED_CMB_HEADERS))),
		title=_("银行流水转换"),
	)


def _validate_import_context(doc, requested_bank):
	if not doc.bank_account or not doc.bank:
		frappe.throw(_("请先选择并保存银行账户。"), title=_("银行流水转换"))
	if requested_bank and requested_bank != doc.bank:
		frappe.throw(_("银行账户已变更，请刷新后重新上传流水。"), title=_("银行流水转换"))
	if frappe.db.get_value("Bank Account", doc.bank_account, "bank") != doc.bank:
		frappe.throw(_("所选银行账户与导入单据银行不一致。"), title=_("银行流水转换"))


def _get_attached_source_file(doc, source_file, doctype=None):
	if not source_file or Path(source_file).suffix.lower() != ".xlsx":
		frappe.throw(_("请上传 .xlsx 格式的银行流水文件。"), title=_("银行流水转换"))

	file_name = frappe.db.get_value(
		"File",
		{"file_url": source_file, "attached_to_doctype": doctype or doc.doctype, "attached_to_name": doc.name},
		"name",
	)
	if not file_name:
		frappe.throw(_("原始流水文件必须作为当前银行对账单导入记录的附件上传。"), title=_("银行流水转换"))
	return frappe.get_doc("File", file_name)


def _find_cmb_header(worksheet):
	for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 50), values_only=True), 1):
		header_indexes = {_normalize_header(value): index for index, value in enumerate(row) if _normalize_header(value)}
		if REQUIRED_CMB_HEADERS.issubset(header_indexes):
			return row_index, header_indexes
	return None, None


def _parse_cmb_rows(worksheet, header_row, header_indexes, bank_account):
	rows = []
	errors = []
	references = set()
	for row_number, values in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
		if not any(value not in (None, "") for value in values):
			continue
		row = _row_as_dict(values, header_indexes)
		try:
			converted = _convert_cmb_row(row, bank_account)
			if converted[5] in references:
				raise ValueError(_("流水号重复：{0}").format(converted[5]))
			references.add(converted[5])
			rows.append(converted)
		except ValueError as exc:
			errors.append(_("第 {0} 行：{1}").format(row_number, exc))

	if errors:
		frappe.throw("<br>".join(errors[:20]), title=_("银行流水转换失败"))
	if not rows:
		frappe.throw(_("未找到可导入的招商银行流水。"), title=_("银行流水转换"))
	return rows


def _row_as_dict(values, header_indexes):
	return {header: values[index] if index < len(values) else None for header, index in header_indexes.items()}


def _convert_cmb_row(row, bank_account):
	transaction_date = _format_date(row.get("交易日"))
	reference = str(row.get("流水号") or "").strip()
	if not reference:
		raise ValueError(_("缺少流水号"))
	deposit = _parse_amount(row.get("贷方金额"), "贷方金额")
	withdrawal = _parse_amount(row.get("借方金额"), "借方金额")
	if deposit and withdrawal:
		raise ValueError(_("借方金额和贷方金额不能同时大于零"))
	if not deposit and not withdrawal:
		raise ValueError(_("借方金额和贷方金额不能同时为空或零"))

	currency = CURRENCY_MAP.get(str(row.get("币种") or "").strip())
	if not currency:
		raise ValueError(_("不支持的币种：{0}").format(row.get("币种") or _("空")))

	description = _build_description(row)
	summary = str(row.get("摘要") or "").strip()
	return [
		transaction_date,
		_format_amount(deposit),
		_format_amount(withdrawal),
		summary,
		description,
		reference,
		bank_account,
		currency,
	]


def _format_date(value):
	if isinstance(value, datetime | date):
		return f"{value.year}/{value.month}/{value.day}"
	try:
		parsed = frappe.utils.getdate(str(value).strip())
	except Exception:
		raise ValueError(_("交易日无效：{0}").format(value or _("空")))
	return f"{parsed.year}/{parsed.month}/{parsed.day}"


def _parse_amount(value, label):
	if value in (None, ""):
		return Decimal("0")
	try:
		amount = Decimal(str(value).replace(",", "").strip())
	except (InvalidOperation, ValueError):
		raise ValueError(_("{0}无效：{1}").format(label, value))
	if amount < 0:
		raise ValueError(_("{0}不能为负数").format(label))
	return amount.normalize() if amount else Decimal("0")


def _format_amount(amount):
	if not amount:
		return ""
	return format(amount, "f").rstrip("0").rstrip(".") if "." in format(amount, "f") else format(amount, "f")


def _build_description(row):
	parts = [str(row.get("摘要") or "").strip()]
	for label, fieldname in (("对方", "收(付)方名称"), ("账号", "收(付)方账号"), ("开户行", "收(付)方开户行名"), ("交易类型", "交易类型")):
		value = str(row.get(fieldname) or "").strip()
		if value:
			parts.append(f"{label}：{value}")
	return "｜".join(part for part in parts if part)


def _normalize_header(value):
	return str(value or "").replace(" ", "").replace("\n", "").strip()


def _get_or_create_converted_file(doc, rows, file_hash, doctype=None, fieldname="import_file"):
	doctype = doctype or doc.doctype
	filename = f"招商银行流水-{file_hash[:12]}.csv"
	existing = frappe.db.get_value(
		"File",
		{"attached_to_doctype": doctype, "attached_to_name": doc.name, "file_name": filename},
		"name",
	)
	if existing:
		return frappe.get_doc("File", existing), True

	buffer = io.StringIO(newline="")
	writer = csv.writer(buffer, lineterminator="\n")
	writer.writerow(STANDARD_HEADERS)
	writer.writerows(rows)
	return (
		save_file(filename, buffer.getvalue().encode("utf-8-sig"), doctype, doc.name, is_private=True, df=fieldname),
		False,
	)
